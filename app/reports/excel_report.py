from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelReport:

    @staticmethod
    def generate(matches, transactions, output_path):

        workbook = Workbook()

        # ===============================
        # Summary Sheet
        # ===============================
        summary = workbook.active
        summary.title = "Summary"

        summary["A1"] = "ReconX Reconciliation Report"
        summary["A1"].font = Font(bold=True, size=16)

        summary["A3"] = "Total Transactions"
        summary["B3"] = len(transactions)

        summary["A4"] = "Matched Pairs"
        summary["B4"] = len(matches)

        summary["A5"] = "Matched Transactions"
        summary["B5"] = len(matches) * 2

        summary["A6"] = "Unmatched Transactions"
        summary["B6"] = len(transactions) - (len(matches) * 2)

        # ===============================
        # Matched Sheet
        # ===============================
        matched_sheet = workbook.create_sheet("Matched")

        matched_sheet.append([
            "Transaction 1",
            "Transaction 2",
            "Score",
            "Reason"
        ])

        for cell in matched_sheet[1]:
            cell.font = Font(bold=True)

        for match in matches:

            matched_sheet.append([
                match["transaction_1"],
                match["transaction_2"],
                match["score"],
                ", ".join(match["reason"])
            ])

        # ===============================
        # Unmatched Sheet
        # ===============================
        unmatched = workbook.create_sheet("Unmatched")

        unmatched.append([
            "Transaction ID",
            "Bank",
            "Amount",
            "Currency",
            "Status"
        ])

        for cell in unmatched[1]:
            cell.font = Font(bold=True)

        for tx in transactions:

            if tx.match_status != "MATCHED":

                unmatched.append([
                    tx.transaction_id,
                    tx.bank_name,
                    float(tx.amount),
                    tx.currency,
                    tx.match_status
                ])

        Path(output_path).parent.mkdir(
            parents=True,
            exist_ok=True
        )

        workbook.save(output_path)

        return output_path